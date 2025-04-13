"""Excel Audit Report Generator"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class ExcelReportGenerator:
    """Excel Report Generator Class"""

    def __init__(self):
        """
        Initialize the generator with an empty workbook.
        """
        self.workbook = None

    def create_workbook(self):
        """
        Create a new workbook.
        """
        self.workbook = Workbook()
        # print("New workbook created.")

    def create_sheet(self, sheet_name,tab_color=None):
        """
        Create a new sheet in the workbook.

        :param sheet_name: The name of the new sheet.
        """
        if self.workbook is None:
            raise ValueError(
                "Workbook not created. Please create a workbook first using CreateWorkbook method.")

        # If the workbook has the default sheet, remove it if this is the first
        # sheet being added.
        if len(self.workbook.sheetnames) == 1 and self.workbook.active.title == "Sheet":
            self.workbook.remove(self.workbook.active)

        ws = self.workbook.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = tab_color
        # print(f"Sheet '{sheet_name}' created.")

    def add_data(self, sheet_name, data):
        """
        :param sheet_name: The name of the sheet where data should be added.
        :param data: A list of dictionaries representing rows of data.
        """
        if self.workbook is None:
            raise ValueError(
                "Workbook not created. Please create a workbook first using CreateWorkbook method.")

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' does not exist. Please create the sheet first using CreateSheet method.")

        sheet = self.workbook[sheet_name]
        if not data:
            return

        # Write headers only if the sheet is empty
        if sheet.max_row == 1 and sheet["A1"].value is None:
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

        # Write data rows
        for row in data:
            sheet.append(list(row.values()))

        # Adjust column width
        headers = list(data[0].keys())
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(cell.value))
                             for cell in sheet[column_letter])
            sheet.column_dimensions[column_letter].width = max_length + 2

        # print(f"Data added to sheet '{sheet_name}'.")

    def save_workbook(self, file_name='report.xlsx'):
        """
        Save the workbook to a file.

        :param file_name: The name of the file to save the workbook as.
        """
        if self.workbook is None:
            raise ValueError(
                "Workbook not created. Please create a workbook first using CreateWorkbook method.")

        self.workbook.save(file_name)
        return file_name
    