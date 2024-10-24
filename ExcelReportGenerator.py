from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class ExcelReportGenerator:
    def __init__(self):
        """
        Initialize the generator with an empty workbook.
        """
        self.workbook = None

    def CreateWorkbook(self):
        """
        Create a new workbook.
        """
        self.workbook = Workbook()
        #print("New workbook created.")

    def CreateSheet(self, sheet_name):
        """
        Create a new sheet in the workbook.
        
        :param sheet_name: The name of the new sheet.
        """
        if self.workbook is None:
            raise ValueError("Workbook not created. Please create a workbook first using CreateWorkbook method.")
        
        # If the workbook has the default sheet, remove it if this is the first sheet being added.
        if len(self.workbook.sheetnames) == 1 and self.workbook.active.title == "Sheet":
            self.workbook.remove(self.workbook.active)
        
        self.workbook.create_sheet(title=sheet_name)
        #print(f"Sheet '{sheet_name}' created.")

    def AddData(self, sheet_name, data):
        """
                
        :param sheet_name: The name of the sheet where data should be added.
        :param data: A list of dictionaries representing rows of data.
        """
        if self.workbook is None:
            raise ValueError("Workbook not created. Please create a workbook first using CreateWorkbook method.")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist. Please create the sheet first using CreateSheet method.")
        
        sheet = self.workbook[sheet_name]
        if not data:
            return
        
        # Write headers only if the sheet is empty
        if sheet.max_row == 1 and sheet["A1"].value is None:
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
           
        # Write data rows
        for row in data:
            sheet.append(list(row.values()))

        # Adjust column width
        headers = list(data[0].keys())
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(cell.value)) for cell in sheet[column_letter])
            sheet.column_dimensions[column_letter].width = max_length + 2
        
        #print(f"Data added to sheet '{sheet_name}'.")

    def SaveWorkbook(self, file_name='report.xlsx'):
        """
        Save the workbook to a file.
        
        :param file_name: The name of the file to save the workbook as.
        """
        if self.workbook is None:
            raise ValueError("Workbook not created. Please create a workbook first using CreateWorkbook method.")
        
        self.workbook.save(file_name)
        print(f"Workbook saved as '{file_name}'.")
'''
# Example usage:
if __name__ == "__main__":
    # Sample data
    summary_data = [
        {"Name": "Alice", "Sales": 3000, "Target": 4000},
        {"Name": "Bob", "Sales": 5000, "Target": 4000},
        {"Name": "Charlie", "Sales": 2000, "Target": 4000}
    ]
    
    additional_data = [
        {"Product": "Widget A", "Price": 50, "Quantity Sold": 60},
        {"Product": "Widget B", "Price": 30, "Quantity Sold": 80},
        {"Product": "Widget C", "Price": 20, "Quantity Sold": 100}
    ]

    # Create a new instance of the generator
    generator = ExcelReportGenerator()

    # Create a new workbook and add data
    generator.CreateWorkbook()
    generator.CreateSheet("Summary")
    generator.AddData("Summary", summary_data)
    
    generator.CreateSheet("Additional Information")
    generator.AddData("Additional Information", additional_data)

    # Save the workbook
    generator.SaveWorkbook('example_report.xlsx')
'''